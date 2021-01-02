import requests
import pandas as pd
from openpyxl import load_workbook
from bs4 import BeautifulSoup


# Global variables
HEADS = {'User-Agent' : 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) \
          AppleWebKit/537.36 (KHTML, like Gecko) \
          Chrome/70.0.3538.110 Safari/537.36'}
FILENAME = 'test.xlsx'
PLAYERLIST = []
# Global variables


def get_response(url):
    '''
    Returns the response of each url.

            Parameters:
                    url (str): A url

            Returns:
                    response (response): A response from requests
    '''
    
    # Catching all exceptions   
    try:
        # url request
        response = requests.get(url, headers=HEADS)
    
    # Handle possible exception
    except AttributeError:
        print('AttributeError') 
    
    return response


def scraping_kader(response): 
    '''
    Starts scraping for players squad.

            Parameters:
                    response (response): A response from requests (Kader-Page)

            Returns:
                    responses (list of response): A list of responses from
                    requests
    '''
    
    # Catching all exceptions  
    try:
        # BeautifulSoup parser
        soup = BeautifulSoup(response.text, 'html.parser')
       
        # Extract the season, the team-id and the team name
        season = response.url.split('/')[-3]
        team_id = response.url.split('/')[-5]
        team_name = soup.find_all('div', 'dataName')[0].span.text
        
        # For debugging
        print(team_id)
        print(season)
        print(team_name)
        
        # Creating an empty list
        player_ids = []
        
        # Extract player-id with BS to combine it later with the PD dataframe
        all_tooltips = soup.find_all('td', class_='hauptlink')
        for item in all_tooltips:
            item = item.find('a', class_='spielprofil_tooltip')
            if item:
               splitstring = item['href'].split('/')
               # The Pandas dataframe uses only every third row later
               player_ids.append(splitstring[-1])
               # Empty row
               player_ids.append('')
               # Empty row
               player_ids.append('')
                
        # Pandas parser
        dfs = pd.read_html(response.text)
        
        # Create an empty list 
        zeilenliste =[] 
        
        # Define and set the column-titles
        columns_titles = ['Verein ID', 'Verein Name', 'Saison', 'Spieler',
                          '#', 'Im Team seit', 'Vertrag bis', 'Marktwert'
                         ]
        dfs[1]=dfs[1].reindex(columns=columns_titles)
        
        # Iterate over every third row
        i = 0
        end = len(dfs[1].index)
        while i < end:
            print(dfs[1].iloc[i, 3])
            print(player_ids[i])
            # Use the information generated above to create a complete row
            dfs[1].iloc[i, 0] = team_id
            dfs[1].iloc[i, 1] = team_name
            dfs[1].iloc[i, 2] = season
            dfs[1].iloc[i, 3] = player_ids[i]
            # Add only the necessary line to the result
            zeilenliste.append(list(dfs[1].iloc[i, :]))
            # Put the player on the separate player-list
            PLAYERLIST.append(int(player_ids[i]))
            # Skip two rows
            i = i + 3
          
        # Select file
        wb = load_workbook(FILENAME)

        # Select worksheet
        ws = wb.worksheets[4]

        # Append all necessary information
        for row_data in zeilenliste:
            ws.append(row_data)

        # Safe file
        wb.save(FILENAME)
        
    # Handle possible exceptions    
    except ImportError:
        print('ImportError')
        
    except ValueError:
        print('ValueError')
        
    except AttributeError:
        print('AttributeError') 
        
    except IndexError:
        print('IndexError')       
        

def label(label_type):
    '''
    Sets the name of the first line per sheet according to the label_type.

            Parameters:
                    label_type (str): A label-type

            Returns:
                    Null - stores all data in the excel file
    '''
    
    if label_type == 'team':
        # Select file
        wb = load_workbook(FILENAME)
    
        # Select worksheet
        ws = wb.worksheets[4]
    
        # Columns labeling
        column_name = ['Verein-ID', 'Verein Name', 'Saison', 'Spieler-ID',
                       'Rückennummer', 'Im Team seit', 'Vertrag bis',
                       'Marktwert'
                       ]
        ws.append(column_name)
    
        # Safe file
        wb.save(FILENAME)

        
def start_single_scraping(club_id, season):
    '''
    Starts scraping with a single club-id in combination with a single season.

            Parameters:
                    club_id (int): A club-id
                    season (int): A season

            Returns:
                    Null - justs starts scraping
    '''
    
    # Creating url by combination of existing information
    url = ('https://www.transfermarkt.de/xxx/kader/verein/'
           + str(club_id) + '/saison_id/' + str(season) + '/plus/1')
    
    # For debugging
    print(url)
    
    # url request 
    response = get_response(url)
    
    # Start scraping for players squad
    scraping_kader(response)
    
    
def start_multi_scraping(club_ids, seasons):
    '''
    Starts scraping with a list of club-ids and a list of seasons.

            Parameters:
                    club_ids (list of int): A list of club-ids
                    seasons (list of int): A list of seasons
            
            Returns:
                    Null - justs starts scraping
    '''
    
    # Use start_single_scraping with every club-id (in the list) ...
    # ... depending on each season (in the list)
    for club_id in club_ids:
        for season in seasons:
            start_single_scraping(club_id, season)
    

##############################################################################
##############################################################################

#label('team')    
#bundesliga = [16, 39, 15, 42, 27, 60, 24, 533, 41, 33, 4, 2, 82, 86, 35, 3, 79, 18, 44, 167, 38, 65, 23, 127, 4795, 105, 23826, 89]
#bundesliga2 = [94, 167, 44, 25, 52, 65, 293, 80, 72, 8, 38, 127, 89, 81, 57, 48, 4795, 10, 24, 35, 23, 129, 30, 2, 83, 3, 254, 109, 105, 23826, 2036, 4, 60, 42, 79, 1557, 269, 41, 187, 108]
#bundesliga3 = [30, 23, 107, 108, 129, 111, 2036, 996, 109, 50, 66, 1, 102, 90, 28, 83, 110, 113, 254, 87, 91, 81, 105, 10, 21, 57, 48, 440, 71, 17, 8, 23826, 52, 64, 269, 25, 92, 2589, 851, 94, 187, 1557, 668, 127, 293, 275, 247, 95, 2, 72, 4795, 85, 1622]
#premierleague = [985, 11, 281, 631, 148, 355, 1181, 289, 762, 31, 512, 984, 405, 29, 164, 337, 1071, 931, 543, 379, 1123, 2288, 1039, 180, 1032, 3008, 603, 873, 1132, 1003, 1010, 989, 641, 1110, 1237, 350]
#primeradivision = [131, 418, 1050, 1049, 714, 13, 621, 3709, 237, 368, 681, 7971, 897, 630, 331, 1084, 2448, 142, 3368, 3302, 150, 367, 16795, 366, 940, 1531, 1533, 993, 472, 1108, 1244, 12321, 5358]
#seriea = [5, 6195, 46, 398, 12, 458, 506, 410, 1038, 252, 1390, 430, 862, 130, 1025, 1627, 1429, 1005, 19, 332, 800, 1387, 6692, 416, 2921, 276, 6574, 1210, 749, 8970, 4102, 4083, 2722, 4171]
#ligue1 = [1082, 583, 273, 1041, 244, 618, 3911, 40, 415, 969, 750, 1158, 1159, 290, 417, 1423, 162, 826, 1162, 11300, 14171, 2969, 1147, 595, 1421, 1095, 995, 855, 347, 1420, 3558, 667, 1416, 1160]
#club_ids = bundesliga+bundesliga2+bundesliga3+premierleague+primeradivision+seriea+ligue1
#club_ids = list(dict.fromkeys(club_ids))
seasons = [2010, 2011, 2012, 2013, 2014, 2015, 2016, 2017, 2018, 2019]
start_multi_scraping([533], seasons)
#print(PLAYERLIST)
#print(list(dict.fromkeys(PLAYERLIST)))
