import requests
import pandas as pd
from openpyxl import load_workbook
from bs4 import BeautifulSoup


# Global variables
HEADS = {'User-Agent' : 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) \
          AppleWebKit/537.36 (KHTML, like Gecko) \
          Chrome/70.0.3538.110 Safari/537.36'}
OTHERS = []
FILENAME = 'test.xlsx'
# Global variables


def scraping_competition(response, club_id, season): 
    '''
    Starts scraping with a single league in combination with a single season.

            Parameters:
                    league (str): A league
                    season (int): A season

            Returns:
                    Null - justs starts scraping
    '''
    
    # Catching all exceptions 
    try:
        # BeautifulSoup parser
        soup = BeautifulSoup(response.text, 'html.parser')
    
        # Initialize two empty lists
        competition = []
        games = []
              
        # Extract the part where to find the competition boxes
        all_links = soup.find_all('div', class_='large-8')
        tabelle = all_links[0].find_all('div', class_='box')

        # Pandas parser
        dfs = pd.read_html(response.text)
    
        # The very first box it the selection of the season - delete
        del tabelle[0]
        
        # Debugging
        print('Tabellemessung 1: ' + str(len(tabelle)))
        
        # Delete the first two boxes when the first box is a graph 
        if tabelle[0].find('h2').text[:24] == 'Tabellenplatzentwicklung':
            del tabelle[0]
            del tabelle[0]
            i = 2
            
            # Debugging (Number of boxes)
            print('Tabellemessung 2: ' + str(len(tabelle)))
            
            # Count each row of the table, each represents one game
            while i <= len(tabelle)+1:
                #print(len(dfs[i].index))
                games.append(len(dfs[i].index))
                i = i +1
        
        # When the first box isn't a graph start directly  
        else:
            i = 1
            
            # Debugging
            print('Tabellemessung 3: ' + str(len(tabelle)))
    
            # Count each row of the table, each represents one game
            while i <= len(tabelle):
                #print(len(dfs[i].index))
                games.append(len(dfs[i].index))
                i = i +1
    
        # Debugging
        print(games)
    
        # The box names represent the competitions
        for item in tabelle:
            item = item.find('a')
            competition.append(item.img['title'])
        
        # Debugging
        print(competition)
    
        # Initialize counters
        league = 0
        cup = 0
        EL = 0
        CL = 0
        other = 0
        
        # Classification or assignment of the competitions
        i = 0
        while i < len(competition):
            
            # League
            if competition[i] == '1. Bundesliga':
                league = league + games[i]
            
            elif competition[i] == 'Relegation 1. Bundesliga':
                league = league + games[i]            
                
            elif competition[i] == '2. Bundesliga':
                league = league + games[i]
                
            elif competition[i] == 'Relegation 2. Bundesliga':
                league = league + games[i]
                
            elif competition[i] == '3. Liga':
                league = league + games[i]
                
            elif competition[i] == 'Regionalliga Bayern':
                league = league + games[i]
                
            elif competition[i] == 'Regionalliga Nord':
                league = league + games[i]
                
            elif competition[i] == 'Aufstiegsrunde zur Regionalliga Südwest':
                league = league + games[i]
                
            elif competition[i] == 'Aufstiegsrunde zur Regionalliga Bayern':
                league = league + games[i]
                
            elif competition[i] == 'Oberliga Baden-Württemberg':
                league = league + games[i]
            
            elif competition[i] == 'Oberliga Rheinland-Pfalz/Saar':
                league = league + games[i]
                
            elif competition[i] == 'Oberliga Niederrhein':
                league = league + games[i]
                
            elif competition[i] == 'Oberliga Niedersachsen':
                league = league + games[i]
                
            elif competition[i] == 'Oberliga Westfalen':
                league = league + games[i]
                
            elif competition[i] == 'NOFV-Oberliga Süd':
                league = league + games[i]

            elif competition[i] == 'Regionalliga Nordost':
                league = league + games[i]
                
            elif competition[i] == 'Regionalliga Südwest':
                league = league + games[i]
                
            elif competition[i] == 'Landesliga Niederrhein - Gruppe 3':
                league = league + games[i]
                   
            elif competition[i] == 'NRW-Liga (bis 11/12)':
                league = league + games[i]
                
            elif competition[i] == 'Regionalliga Süd (bis 11/12)':
                league = league + games[i]
                
            elif competition[i] == 'Regionalliga Nord (bis 11/12)':
                league = league + games[i]
                
            elif competition[i] == 'Regionalliga West':
                league = league + games[i]
            
            elif competition[i] == 'Regionalliga West (bis 11/12)':
                league = league + games[i]
                
            elif competition[i] == 'Aufstiegsrunde zur 3. Liga':
                league = league + games[i]
                
            elif competition[i] == 'Oberliga Rheinland-Pfalz/Saar':
                league = league + games[i]
                
            elif competition[i] == 'Aufstiegsrunde zur Regionalliga West':
                league = league + games[i]
                
            elif competition[i] == 'NRW-Liga (bis 11/12)':
                league = league + games[i]
    
            elif competition[i] == 'LaLiga':
                league = league + games[i]
                
            elif competition[i] == 'Aufstiegs-Playoff LaLiga':
                league = league + games[i]

            elif competition[i] == 'LaLiga2':
                league = league + games[i]
                
            elif competition[i] == 'Aufstiegs-Playoff LaLiga2':
                league = league + games[i]
                
            elif competition[i] == 'Segunda División B - Grupo III':
                league = league + games[i]
                
            elif competition[i] == 'Segunda División B - Grupo I':
                league = league + games[i]
                
            elif competition[i] == 'Segunda División B - Grupo II':
                league = league + games[i]
                
            elif competition[i] == 'Segunda División B - Grupo IV':
                league = league + games[i]
                
            elif competition[i] == 'Premier League':
                league = league + games[i]
                
            elif competition[i] == 'League One':
                league = league + games[i]
                
            elif competition[i] == 'League One Playoffs':
                league = league + games[i]
                
            elif competition[i] == 'League Two':
                league = league + games[i]
                
            elif competition[i] == 'League Two Playoffs':
                league = league + games[i]
                
            elif competition[i] == 'Championship':
                league = league + games[i]
                
            elif competition[i] == 'Championship Playoffs':
                league = league + games[i]
                
            elif competition[i] == 'Serie A':
                league = league + games[i]
    
            elif competition[i] == 'Serie B':
                league = league + games[i]
                
            elif competition[i] == 'Serie B Playoff':
                league = league + games[i]
                
            elif competition[i] == 'Serie D - Girone I':
                league = league + games[i]
                
            elif competition[i] == 'Serie C - Girone B':
                league = league + games[i]
                
            elif competition[i] == 'Serie C - Girone C':
                league = league + games[i]
                
            elif competition[i] == 'Serie D - Girone F':
                league = league + games[i]
                
            elif competition[i] == 'Play-off Serie C':
                league = league + games[i]
                
            elif competition[i] == 'Serie C - Girone A':
                league = league + games[i]
                
            elif competition[i] == 'Serie D - Girone E':
                league = league + games[i]
                
            elif competition[i] == 'Serie D - Girone D':
                league = league + games[i]
                
            elif competition[i] == 'Serie B Play-out':
                league = league + games[i]
                
            elif competition[i] == 'Play-out Serie C':
                league = league + games[i]
    
            elif competition[i] == 'Ligue 1':
                league = league + games[i]
                
            elif competition[i] == 'Ligue 2':
                league = league + games[i]
                
            elif competition[i] == 'Relegation Ligue 1':
                league = league + games[i]
                
            elif competition[i] == 'Championnat National':
                league = league + games[i]
                
            elif competition[i] == 'Championnat National 2 - Groupe A':
                league = league + games[i]
                
            elif competition[i] == 'Relegation Ligue 2':
                league = league + games[i]
                        
            # Cup
            elif competition[i] == 'DFB-Pokal':
                cup = cup + games[i]
            
            elif competition[i] == 'FA Cup':
                cup = cup + games[i]
                
            elif competition[i] == 'Coupe de France':
                cup = cup + games[i]
                
            elif competition[i] == 'Coupe de la Ligue':
                cup = cup + games[i]
                            
            elif competition[i] == 'Copa del Rey':
                cup = cup + games[i]
                            
            elif competition[i] == 'Coppa Italia':
                cup = cup + games[i]
    
            elif competition[i] == 'EFL Cup':
                cup = cup + games[i]
   
            elif competition[i] == 'EFL Trophy':
                cup = cup + games[i] 
   
            elif competition[i] == 'Qualifikationsspiel zum DFB-Pokal (FLVW)':
                cup = cup + games[i]             
   
            elif competition[i] == 'Landespokal Südwest':
                cup = cup + games[i]

            elif competition[i] == 'Landespokal Bayern':
                cup = cup + games[i]
                
            elif competition[i] == 'Landespokal Westfalen':
                cup = cup + games[i]
                
            elif competition[i] == 'Landespokal Hessen':
                cup = cup + games[i]
                
            elif competition[i] == 'Landespokal Baden':
                cup = cup + games[i]     
                
            elif competition[i] == 'Landespokal Brandenburg':
                cup = cup + games[i]                
                
            elif competition[i] == 'Landespokal Mittelrhein':
                cup = cup + games[i]
                
            elif competition[i] == 'Landespokal Niederrhein':
                cup = cup + games[i]  
                
            elif competition[i] == 'Landespokal Württemberg':
                cup = cup + games[i]  
                
            elif competition[i] == 'Landespokal Mecklenburg-Vorpommern':
                cup = cup + games[i]
                
            elif competition[i] == 'Landespokal Sachsen':
                cup = cup + games[i]  
                
            elif competition[i] == 'Landespokal Thüringen':
                cup = cup + games[i]  
              
            elif competition[i] == 'Landespokal Rheinland':
                cup = cup + games[i]
                
            elif competition[i] == 'Landespokal Saarland':
                cup = cup + games[i]     
                        
            elif competition[i] == 'Landespokal Sachsen-Anhalt':
                cup = cup + games[i]  
                
            elif competition[i] == 'Landespokal Schleswig-Holstein':
                cup = cup + games[i] 
              
            elif competition[i]\
                == 'Landespokal Niedersachsen (3. und 4. Liga)':
                cup = cup + games[i]  
                
            elif competition[i] == 'Niedersachsenpokal (bis 17/18)':
                cup = cup + games[i] 
                
            elif competition[i]\
                == 'Landespokal Niedersachsen (3. und 4. Liga)':
                cup = cup + games[i]            
  
            elif competition[i] == 'Community Shield':
                other = other + games[i]
                  
            # International                
            elif competition[i] == 'Europa League':
                EL = EL + games[i]
                
            elif competition[i] == 'Europa League Qualifikation':
                EL = EL + games[i]
    
            elif competition[i] == 'UEFA Champions League':
                CL = CL + games[i]
                
            elif competition[i] == 'UEFA Champions League-Qualifikation':
                CL = CL + games[i]
            
            # Other    
            else:
                other = other + games[i]
                OTHERS.append(competition[i])
    
            i = i + 1
    
        # Debugging (Result check)
        print('Liga ' + str(league))
        print('Pokal ' + str(cup))
        print('EL ' + str(EL))
        print('CL ' + str(CL))
        print('Andere ' + str(other))
        
        # Select file
        wb = load_workbook(FILENAME)
    
        # Select worksheet
        ws = wb.worksheets[5]
    
        # Append all necessary information
        ws.append([club_id, '', season, league, EL, CL, cup, other])
    
        # Safe file  
        wb.save(FILENAME)
    
    # Handle possible exceptions
    except IndexError:
        print('IndexError')             
        
        
def get_response(url):
    '''
    Returns the response of each url.

            Parameters:
                    url (str): A url

            Returns:
                    response (response): A response from requests
    '''
    
    # Catching all exceptions 
    try:
        # url request  
        response = requests.get(url, headers=HEADS)
     
    # Handle possible exceptions    
    except AttributeError:
        print('AttributeError') 
    
    return response


def start_single_scraping(club_id, season):
    '''
    Starts scraping with a single club-id in combination with a single season.

            Parameters:
                    club_id (int): A club-id
                    season (int): A season

            Returns:
                    Null - justs starts scraping
    '''
    
    # Debugging
    print('Team: ' + str(club_id) + '        Saison: ' + str(season))
    
    # Creating url by combination of existing information
    url = ('https://www.transfermarkt.de/xxx/spielplan/verein/'
           + str(club_id) + '/plus/0?saison_id=' + str(season)
           )
    
    # Start scraping competitions
    scraping_competition(get_response(url), club_id, season)
    
    
def start_multi_scraping(club_ids, seasons):
    '''
    Starts scraping with a list of leagues and a list of seasons.

            Parameters:
                    leagues (list of str): A list of club-ids
                    seasons (list of int): A list of seasons
            
            Returns:
                    Null - justs starts scraping
    '''
    
    # Use start_single_scraping with every club-id (in the list) ...
    # ... depending on each season (in the list)
    for club_id in club_ids:
        for season in seasons:
            start_single_scraping(club_id, season)
            

def label(typ):
    '''
    Sets the name of the first line per sheet according to the label_type.

            Parameters:
                    label_type (str): A label-type

            Returns:
                    Null - stores all data in the excel file
    '''
    
    if typ == 'team':
        # Select file
        wb = load_workbook(FILENAME)
    
        # Select worksheet
        ws = wb.worksheets[5]
    
        # Columns labeling
        column_name = ['Verein ID', 'Verein Name', 'Saison', 'Liga',
                       'EL', 'CL', 'Pokal', 'Sonstige'
                       ]
        ws.append(column_name)
    
        # Safe file 
        wb.save(FILENAME)
  
 
##############################################################################
##############################################################################   

label('team')
#bundesliga = [16, 39, 15, 42, 27, 60, 24, 533, 41, 33, 4, 2, 82, 86, 35, 3, 79, 18, 44, 167, 38, 65, 23, 127, 4795, 105, 23826, 89]
#bundesliga2 = [94, 167, 44, 25, 52, 65, 293, 80, 72, 8, 38, 127, 89, 81, 57, 48, 4795, 10, 24, 35, 23, 129, 30, 2, 83, 3, 254, 109, 105, 23826, 2036, 4, 60, 42, 79, 1557, 269, 41, 187, 108]
#bundesliga3 = [30, 23, 107, 108, 129, 111, 2036, 996, 109, 50, 66, 1, 102, 90, 28, 83, 110, 113, 254, 87, 91, 81, 105, 10, 21, 57, 48, 440, 71, 17, 8, 23826, 52, 64, 269, 25, 92, 2589, 851, 94, 187, 1557, 668, 127, 293, 275, 247, 95, 2, 72, 4795, 85, 1622]
#premierleague = [985, 11, 281, 631, 148, 355, 1181, 289, 762, 31, 512, 984, 405, 29, 164, 337, 1071, 931, 543, 379, 1123, 2288, 1039, 180, 1032, 3008, 603, 873, 1132, 1003, 1010, 989, 641, 1110, 1237, 350]
#primeradivision = [131, 418, 1050, 1049, 714, 13, 621, 3709, 237, 368, 681, 7971, 897, 630, 331, 1084, 2448, 142, 3368, 3302, 150, 367, 16795, 366, 940, 1531, 1533, 993, 472, 1108, 1244, 12321, 5358]
#seriea = [5, 6195, 46, 398, 12, 458, 506, 410, 1038, 252, 1390, 430, 862, 130, 1025, 1627, 1429, 1005, 19, 332, 800, 1387, 6692, 416, 2921, 276, 6574, 1210, 749, 8970, 4102, 4083, 2722, 4171]
#ligue1 = [1082, 583, 273, 1041, 244, 618, 3911, 40, 415, 969, 750, 1158, 1159, 290, 417, 1423, 162, 826, 1162, 11300, 14171, 2969, 1147, 595, 1421, 1095, 995, 855, 347, 1420, 3558, 667, 1416, 1160]
#club_ids = bundesliga+bundesliga2+bundesliga3+premierleague+primeradivision+seriea+ligue1
#club_ids = list(dict.fromkeys(club_ids))   
seasons = [2010]
club_ids = [533]
start_multi_scraping(club_ids, seasons)
#print(list(dict.fromkeys(OTHERS)))
