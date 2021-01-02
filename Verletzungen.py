import requests
import time
import pandas as pd
from openpyxl import load_workbook
from bs4 import BeautifulSoup


# Global variables
HEADS = {'User-Agent' : 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) \
          AppleWebKit/537.36 (KHTML, like Gecko) \
          Chrome/70.0.3538.110 Safari/537.36'}
FILENAME = 'test.xlsx'
PROFIL_COLUMN_NAMES = ['Spieler-ID', 'Nachname', 'Vorname',
                       'Geburtsdatum', 'Geburtsort', 'Größe',
                       'Nationalität', 'Position', 'Fuß',
                       'Spielerberater', 'Aktueller Verein',
                       'Im Team seit', 'Vertrag bis',
                       'Letzte Vertragsverlängerung', 'Ausrüster',
                       'Social Media', 'Todestag', 'Alter',
                       'Name im Heimatland', 'Weitere Attribute'
                       ]
# Global variables


def get_response_verletzungshistorie(player_id):
    '''
    Returns the response of the Verletzungshistorien-Page.

            Parameters:
                    player_id (int): A player-id

            Returns:
                    response (response): A response from requests
    '''
    
    # Creates the player_url from base und player_id
    url = 'https://www.transfermarkt.de/xxx/verletzungen/spieler/'
    player_url = url + str(player_id)

    # Catching all exceptions 
    try:
        # url request
        response = requests.get(player_url, headers=HEADS)
        
        # Logging / Bugfixing
        print('   "Verletzungshistorie"' + '   Status-code: ' 
              + str(response.status_code)
              )
        
    except AttributeError:
        # Logging / Bugfixing
        print('player-id: ' + str(player_id) + ' AttributeError') 
    
    return response


def get_response(url):
    '''
    Returns the response of each url.

            Parameters:
                    url (str): A url

            Returns:
                    response (response): A response from requests
    '''
    
    # Catching all exceptions 
    try:
        # url request
        response = requests.get(url, headers=HEADS)
    
    # Handle possible exception    
    except AttributeError:
        print('AttributeError') 
    
    return response


def get_response_profil(profil_link):
    '''
    Returns the response of the Profil-Page url.

            Parameters:
                    profil_link (str): A url

            Returns:
                    response (response): A response from requests
    '''
    
    # Catching all exceptions 
    try:
        # url request 
        response = requests.get(profil_link, headers=HEADS)
        print('   "Profil"             ' + '   Status-Code: '
              + str(response.status_code)
              )
     
    # Handle possible exception   
    except AttributeError:
        print('"Profil" AttributeError') 
    
    return response
    

def extract_profil_link(response):
    '''
    Returns the profil_link based on Verletzungshistorien-Page.

            Parameters:
                    response (response): A response from requests 
                    Verletzungshistorien-Page)

            Returns:
                    profil_link (str): A String to a Profil-Page
    '''
    
    # BeautifulSoup parser
    soup = BeautifulSoup(response.text, "html.parser")
    
    # Creating link by using BeautifulSoup filter
    profil_link = ('https://www.transfermarkt.de'
                  + str(soup.find_all('a', 'megamenu')[0]['href'])
                  )
    
    return profil_link


def count_subpages_verletzungshistorie(response):
    '''
    Returns the number of subpages based on Verletzungshistorien-Page.

            Parameters:
                    response (response): A response from requests 
                    (Verletzungshistorien-Seite)

            Returns:
                    subpages (int): A number of subpages
    '''    
    
    # BeautifulSoup parser
    soup = BeautifulSoup(response.text, "html.parser")
    
    # BeautifulSoup filter
    pages = soup.find_all("li", "page")
    
    # Returns the number if there is more than 1 subpage
    if len(pages) > 0:
        return len(pages)
    else:
        return 1
    

def get_all_verletzungshistorien_responses(response):
    '''
    Returns the responses of each suburl.

            Parameters:
                    response (response): A response from requests 
                    (Verletzungshistorien-Page)

            Returns:
                    responses (list of response): A list of responses from 
                    requests
    '''
    
    # Get the number of subpages based on Verletzungshistorien-Page
    subpages = count_subpages_verletzungshistorie(response)
    
    # Safe the current url building the urls from the subpages
    url = response.url
    
    # Create an empty list
    responses = {}
    
    # Response 1 is already available
    responses[0] = response
    
    # Create and return all necessary subpages
    if subpages == 1:
        return responses
    else:
        i = 2
        while i <= subpages:
            sub_url = url + '/page/' + str(i)
            sub_response = get_response(sub_url)
            responses[i-1] = sub_response
            i = i + 1
        return responses
   
    
def scraping_all_verletzungshistorien(response, player_id):
    '''
    Starts scraping of all Verletzungshistorien-Pages by checking the subpages.

            Parameters:
                    response (response): A response from requests 
                    (Verletzungshistorien-Page)
                    player_id (int): A player-id

            Returns:
                    responses (list of response): A list of responses from
                    requests
    '''
    
    # Getting all responses of all subpages
    responses = get_all_verletzungshistorien_responses(response)
    
    # Scrapes every subpage
    for i in responses:
        scraping_verletzungshistorie(responses[i], player_id)

    
def scraping_verletzungshistorie(response, player_id): 
    '''
    Scrapes a Verletzungshistorien-Page.

            Parameters:
                    response (response): A response from requests
                    (Verletzungshistorien-Page)
                    player_id (int): A player-id

            Returns:
                    Null - stores all data in the excel file
    '''
    
    # Catching all exceptions
    try:
        # Pandas parser
        dfs = pd.read_html(response.text)

        # Create an empty list 
        row_list =[] 
    
        # Insert player-id into the row
        dfs[0].insert(0, 'player_id', player_id)
  
        # Iterate over each row 
        for i in range((dfs[0].shape[0])): 
            # Add the current row
            row_list.append(list(dfs[0].iloc[i, :])) 
  
        # Select file
        wb = load_workbook(FILENAME)

        # Select worksheet
        ws = wb.worksheets[2]

        # Append all injuries (each in one row)
        for row_data in row_list:
            ws.append(row_data)

        # Safe file
        wb.save(FILENAME)
    
    # Handle possible exceptions
    except ImportError:
        print(str(player_id) + ' ImportError')
        
    except ValueError:
        print(str(player_id) + ' ValueError')
        
    except AttributeError:
        print(str(player_id) + ' AttributeError') 


def scraping_profil(response, player_id):    
    '''
    Scrapes a Profil-Page.

            Parameters:
                    response (response): A response from requests (Profil-Page)
                    player_id (int): A player-id

            Returns:
                    Null - stores all data in the excel file
    '''
    
    try:
        # Pandas parser
        dfs = pd.read_html(response.text)
        
        # Create empty dataframe and adddefined column names
        df = pd.DataFrame({}, 
                 index = ['0'],
                 columns = PROFIL_COLUMN_NAMES)
        
        # BeautifulSoup parser
        soup = BeautifulSoup(response.text, "html.parser")
        
        # BeautifulSoup filter and extract first name and last name 
        soupname = soup.find('h1') 
        name = str(soupname.text)
        lastname = str(soupname.b.text)
        firstname = name.replace(lastname, '')
        firstname = firstname[:-1]
        
        # Set existing values
        df.at[df.index[0],df.columns[0]] = player_id
        df.at[df.index[0],df.columns[1]] = lastname
        df.at[df.index[0],df.columns[2]] = firstname
        
        # Iterate over each row anch check if we use it in our file
        for i in range((dfs[0].shape[0])): 
            if str(dfs[0].at[dfs[0].index[i],dfs[0].\
               columns[0]]).replace(':', '') == 'Geburtsdatum':
                df.at[df.index[0],df.columns[3]]\
                = dfs[0].at[dfs[0].index[i],dfs[0].columns[1]]
                
            if str(dfs[0].at[dfs[0].index[i],dfs[0].\
               columns[0]]).replace(':', '') == 'Geburtsort':
                df.at[df.index[0],df.columns[4]]\
                = dfs[0].at[dfs[0].index[i],dfs[0].columns[1]]
                
            if str(dfs[0].at[dfs[0].index[i],dfs[0].\
               columns[0]]).replace(':', '') == 'Größe':
                df.at[df.index[0],df.columns[5]]\
                = dfs[0].at[dfs[0].index[i],dfs[0].columns[1]]
   
            if str(dfs[0].at[dfs[0].index[i],dfs[0].\
               columns[0]]).replace(':', '') == 'Nationalität':
                df.at[df.index[0],df.columns[6]]\
                = dfs[0].at[dfs[0].index[i],dfs[0].columns[1]]
                
            if str(dfs[0].at[dfs[0].index[i],dfs[0].\
               columns[0]]).replace(':', '') == 'Position':
                df.at[df.index[0],df.columns[7]]\
                = dfs[0].at[dfs[0].index[i],dfs[0].columns[1]]
                
            if str(dfs[0].at[dfs[0].index[i],dfs[0].\
               columns[0]]).replace(':', '') == 'Fuß':
                df.at[df.index[0],df.columns[8]]\
                = dfs[0].at[dfs[0].index[i],dfs[0].columns[1]]
                
            if str(dfs[0].at[dfs[0].index[i],dfs[0].\
               columns[0]]).replace(':', '') == 'Spielerberater':
                df.at[df.index[0],df.columns[9]]\
                = dfs[0].at[dfs[0].index[i],dfs[0].columns[1]]           
 
            if str(dfs[0].at[dfs[0].index[i],dfs[0].\
               columns[0]]).replace(':', '') == 'Aktueller Verein':
                df.at[df.index[0],df.columns[10]]\
                = dfs[0].at[dfs[0].index[i],dfs[0].columns[1]]     
                
            if str(dfs[0].at[dfs[0].index[i],dfs[0].\
               columns[0]]).replace(':', '') == 'Im Team seit':
                df.at[df.index[0],df.columns[11]]\
                = dfs[0].at[dfs[0].index[i],dfs[0].columns[1]]  
            
            if str(dfs[0].at[dfs[0].index[i],dfs[0].\
               columns[0]]).replace(':', '') == 'Vertrag bis':
                df.at[df.index[0],df.columns[12]]\
                = dfs[0].at[dfs[0].index[i],dfs[0].columns[1]]  

            if str(dfs[0].at[dfs[0].index[i],dfs[0].\
               columns[0]]).replace(':', '') == 'Letzte Vertragsverlängerung':
                df.at[df.index[0],df.columns[13]]\
                = dfs[0].at[dfs[0].index[i],dfs[0].columns[1]]                  
 
            if str(dfs[0].at[dfs[0].index[i],dfs[0].\
               columns[0]]).replace(':', '') == 'Ausrüster':
                df.at[df.index[0],df.columns[14]]\
                = dfs[0].at[dfs[0].index[i],dfs[0].columns[1]]    
                
            if str(dfs[0].at[dfs[0].index[i],dfs[0].\
               columns[0]]).replace(':', '') == 'Social Media':
                df.at[df.index[0],df.columns[15]]\
                = dfs[0].at[dfs[0].index[i],dfs[0].columns[1]] 
                
            if str(dfs[0].at[dfs[0].index[i],dfs[0].\
               columns[0]]).replace(':', '') == 'Todestag':
                df.at[df.index[0],df.columns[16]]\
                = dfs[0].at[dfs[0].index[i],dfs[0].columns[1]]
                
            if str(dfs[0].at[dfs[0].index[i],dfs[0].\
               columns[0]]).replace(':', '') == 'Alter':
                df.at[df.index[0],df.columns[17]]\
                    = dfs[0].at[dfs[0].index[i],dfs[0].columns[1]] 
   
            if str(dfs[0].at[dfs[0].index[i],dfs[0].\
               columns[0]]).replace(':', '') == 'Name im Heimatland':
                df.at[df.index[0],df.columns[18]]\
                    = dfs[0].at[dfs[0].index[i],dfs[0].columns[1]] 
            
            # If there is no predefined value put it in a special column ...
            # ... for debugging or later extension
            if str(dfs[0].at[dfs[0].index[i],dfs[0].\
               columns[0]]).replace(':', '') not in PROFIL_COLUMN_NAMES:
                df.at[df.index[0],df.columns[19]]\
                    = str(dfs[0].at[dfs[0].index[i],\
                                    dfs[0].columns[0]]).replace(':', '')  
        
        # Select file
        wb = load_workbook(FILENAME)

        # Select worksheet
        ws = wb.worksheets[1]
        
        # Append all of the information
        ws.append(list(df.iloc[0, :]))

        # Safe file
        wb.save(FILENAME)
    
    # Handle possible exceptions
    except ImportError:
        print(str(player_id) + ' ImportError')
        
    except ValueError:
        print(str(player_id) + ' ValueError')
        
    except AttributeError:
        print(str(player_id) + ' AttributeError') 


def label(label_type):
    '''
    Sets the name of the first line per sheet according to the label_type.

            Parameters:
                    label_type (str): A label-type

            Returns:
                    Null - stores all data in the excel file
    '''
    
    if label_type == 'profil':
        # Select file
        wb = load_workbook(FILENAME)
    
        # Select worksheet
        ws = wb.worksheets[1]
    
        # Columns labeling
        ws.append(PROFIL_COLUMN_NAMES)
    
        # Safe file
        wb.save(FILENAME)
    
    if label_type == 'verletzungshistorie':
        # Select file
        wb = load_workbook(FILENAME)
    
        # Select worksheet
        ws = wb.worksheets[2]
    
        # Columns labeling
        column_name = ['Spieler-ID', 'Saison', 'Verletzung', 'von',
                       'bis', 'Tage', 'Verpasste Spiele'
                       ]
        ws.append(column_name)
    
        # Safe file
        wb.save(FILENAME)
       
    if label_type == 'log':
        # Select file
        wb = load_workbook(FILENAME)
    
        # Select worksheet
        ws = wb.worksheets[0]
    
        # Columns labeling
        column_name = ['Spieler-ID', 'valid player',
                       'status_code Verletzungshistorie',
                       'status_code Profil', 'Timestamp'
                       ]
        ws.append(column_name)
    
        # Safe file
        wb.save(FILENAME) 


def label_columns():
    '''
    Calls the label function with different label_types.

            Parameters:
                    Null - is only used for clarity

            Returns:
                    Null - stores all data in the excel file
    '''
    
    label('log')
    label('profil')
    label('verletzungshistorie')
    
    
def check_valid_player(response):
    '''
    Checks if the Verletzungshistorien-Page exists. 

            Parameters:
                    response (response): A response from requests
                    (Verletzungshistorien-Page)

            Returns:
                    result (boolean): True if Verletzungshistorien-Page exists
    '''
    
    try:
        # BeautifulSoup parser
        soup = BeautifulSoup(response.text, "html.parser")
        
        # BeautifulSoup filter
        titel = soup.title
        
        # True if Verletzungshistorie can be found in title
        result = 'Verletzungshistorie' in str(titel.text)
        
        return result
    
    # Handle possible exceptions and return False
    except AttributeError:
        print('AttributeError')
        return False


def logging_valid_player(player_id, response_verletzungshistorie,
                         response_profil):
    '''
    Safes Logging Data for a valid player.

            Parameters:
                    player_id (int): A player-id
                    response_verletzungshistorie (response): A response from
                    requests (Verletzungshistorien-Page)
                    response_profil (response): A response from requests
                    (Profil-Page)

            Returns:
                    Null - stores all data in the excel file
    '''
       
    # Select file
    wb = load_workbook(FILENAME)

    # Select worksheet
    ws = wb.worksheets[0]

    # Add the row data
    row_data = [player_id,
               True,
               response_verletzungshistorie.status_code,
               response_profil.status_code,
               time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())]
    ws.append(row_data)
    
    # Safe file
    wb.save(FILENAME)
    
        
def logging_none_valid_player(player_id):
    '''
    Safes Logging Data for a none valid player.

            Parameters:
                    player_id (int): A player-id

            Returns:
                    Null - stores all data in the excel file
    '''
    
    # Select file
    wb = load_workbook(FILENAME)

    # Select worksheet
    ws = wb.worksheets[0]

    # Add the row data
    row_data = [player_id,
               False,
               '',
               '',
               time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())]
    ws.append(row_data)
    
    # Safe file
    wb.save(FILENAME)
    
        
def start_multi_scraping(start, end):
    '''
    Starts scraping with an area of player-ids.

            Parameters:
                    start (int): A player-id
                    end (int): A player-id

            Returns:
                    Null - justs starts scraping
    '''
    
    # Use start_single_scraping with every player-id from the defined range
    i = start
    while i <= end:
        print('Starte Scraping: ' + str(i))
        start_single_scraping(i)
        i = i + 1
        
        
def start_single_scraping(player_id):
    '''
    Starts scraping with a single player-id.

            Parameters:
                    player_id (int): A player-id

            Returns:
                    Null - justs starts scraping
    '''
    
    # Get the respone of the Verletzungshistorien-Page to check the ...
    # ... availability of the player
    response_verletzungshistorie = get_response_verletzungshistorie(player_id)
    valid_player = check_valid_player(response_verletzungshistorie)

    # For all valid
    if valid_player:
        # Generate the exact profile-link
        profil_link = extract_profil_link(response_verletzungshistorie)
        # Get the response of the profile-page
        response_profil = get_response_profil(profil_link)
        # Use the verletzungshistorien-response to scrape the injuries
        scraping_all_verletzungshistorien(response_verletzungshistorie,
                                          player_id)
        # Scrapte the profile-information
        scraping_profil(response_profil, player_id)
        # Write the log for debugging
        logging_valid_player(player_id, response_verletzungshistorie, 
                             response_profil)

    # For all non valid players write the log directly
    else:
        logging_none_valid_player(player_id)
        
        
def start_list_scraping(player_ids):
    '''
    Starts scraping with a list of player-ids.

            Parameters:
                    player_ids (list of int): A list of player-ids
            
            Returns:
                    Null - justs starts scraping
    '''
    
    # Use start_single_scraping with every player-id from the list
    i = 1
    for player_id in player_ids:
        print('player-id: ' + str(player_id) + '       '
              + str(i) + ' von ' + str(len(player_ids))
              )
        start_single_scraping(player_id)
        i = i + 1
        

##############################################################################
##############################################################################

#print(get_response('https://www.transfermarkt.de/xxx/verletzungen/spieler/4360/page/3'))
#print('ZEITMESSUNG 1 ' + str(datetime.now()))
#print(read_players())
#print(PLAYERLIST)
label_columns()
#start_multi_scraping(1, 100)
start_single_scraping(72333)
#player_ids = [524, 627, 800, 136098, 8175, 74177, 80351, 212400, 24303, 157491, 163177, 168452, 105622, 41176, 42262, 109576, 47290, 47710, 114923, 124673, 62710, 62731, 62744, 129129]
#start_list_scraping(player_ids)