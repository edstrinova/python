import requests
import pandas as pd
from openpyxl import load_workbook
from bs4 import BeautifulSoup


# Global variables
HEADS = {'User-Agent' : 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) \
          AppleWebKit/537.36 (KHTML, like Gecko) \
          Chrome/70.0.3538.110 Safari/537.36'}
FILENAME = 'test.xlsx'
CLUBLIST = []
# Global variables


def get_response(url):
    '''
    Returns the response of each url.

            Parameters:
                    url (str): A url

            Returns:
                    response (response): A response from requests
    '''
    
    # Catching all exceptions 
    try:
        # url request
        response = requests.get(url, headers=HEADS)
        
    # Handle possible exception    
    except AttributeError:
        print('AttributeError') 
    
    return response

    
def build_url(league, season):
    '''
    Returns a urls for the first half and the second half of the season.

            Parameters:
                    league (str): A name of a league
                    season (int): A year as number

            Returns:
                    urls (str collection): A collection of urls
    '''
    
    # Distinguish by league how many matchdays there are
    if league == '1-bundesliga':
        url = ('https://www.transfermarkt.de/1-bundesliga/'
               'spieltagtabelle/wettbewerb/L1?saison_id=')
        round_count = '&spieltag=17'
        backround_count = '&spieltag=34'
    if league == '2-bundesliga':
        url = ('https://www.transfermarkt.de/2-bundesliga/'
               'spieltagtabelle/wettbewerb/L2?saison_id=')
        round_count = '&spieltag=17'
        backround_count = '&spieltag=34'
    if league == '3-liga':
        url = ('https://www.transfermarkt.de/3-liga/'
               'spieltagtabelle/wettbewerb/L3?saison_id=')
        round_count = '&spieltag=19'
        backround_count = '&spieltag=38'
    if league == 'premier-league':
        url = ('https://www.transfermarkt.de/premier-league/'
               'spieltagtabelle/wettbewerb/GB1?saison_id=')
        round_count = '&spieltag=19'
        backround_count = '&spieltag=38'
    if league == 'primera-division':
        url = ('https://www.transfermarkt.de/primera-division/'
               'spieltagtabelle/wettbewerb/ES1?saison_id=')
        round_count = '&spieltag=19'
        backround_count = '&spieltag=38'
    if league == 'serie-a':
        url = ('https://www.transfermarkt.de/serie-a/'
               'spieltagtabelle/wettbewerb/IT1?saison_id=')
        round_count = '&spieltag=19'
        backround_count = '&spieltag=38'
    if league == 'ligue-1':
        url = ('https://www.transfermarkt.de/ligue-1/'
               'spieltagtabelle/wettbewerb/FR1?saison_id=')
        round_count = '&spieltag=19'
        backround_count = '&spieltag=38'
        # In season 2019 there were only 28 matchdays in France
        if season == 2019:
            backround_count = '&spieltag=28'     
        
    # Creating urls by combination of existing information    
    round_url = url + str(season) + round_count
    backround_url = url + str(season) + backround_count
    
    # Using a dictionary collection to adress the urls
    urls = dict();
    urls['round_url'] = round_url
    urls['backround_url'] = backround_url
    
    print(urls)
    
    return urls


def start_single_scraping(league, season):
    '''
    Starts scraping with a single league in combination with a single season.

            Parameters:
                    league (str): A league
                    season (int): A season

            Returns:
                    Null - justs starts scraping
    '''
    
    # Adress the urls for first half and the second half of the season
    urls = build_url(league, season)
    
    ## Starting with first half of the season
    
    # url request
    response = get_response(urls['round_url'])
    
    # BeautifulSoup parser
    soup = BeautifulSoup(response.text, 'html.parser')
             
    # Extract the part where to find the club ids on the website
    all_links = soup.find_all('div', class_='large-8')
    tabelle = all_links[0].find_all('div', class_='box')
    all_zentriert = tabelle[2].find_all('td', class_='zentriert')

    # Creating an empty list
    club_ids = []

    # Collect all club ids
    for item in all_zentriert:
        item = item.find('a', class_='vereinprofil_tooltip')
        if item:
            club_ids.append(item['id'])

    # Pandas parser
    dfs = pd.read_html(response.text)
    
    # Creating an empty list
    row =[] 
    
    # Define the column titles for reindexing
    columns_titles = ['Season', 'League', '#', 'Verein', 'Verein.1',
                      'Unnamed: 3', 'G', 'U', 'V', 'Tore', '+/-', 'Pkt.'
                      ]
    dfs[4]=dfs[4].reindex(columns=columns_titles)
    
    # Use known information to create a complete row
    for i in range((dfs[4].shape[0])): 
        dfs[4].iloc[i, 0] = str(season)
        dfs[4].iloc[i, 1] = league
        dfs[4].iloc[i, 3] = club_ids[i]
        CLUBLIST.append(int(club_ids[i]))
        row.append(list(dfs[4].iloc[i, :])) 
  
    # Select file
    wb = load_workbook(FILENAME)

    # Select worksheet
    ws = wb.worksheets[3]

    # Append all necessary information
    for row_data in row:
        ws.append(row_data)

    # Safe file   
    wb.save(FILENAME) 
    
    ## Starting with first half of the season
    
    # url request
    response = get_response(urls['backround_url'])
    
    # BeautifulSoup parser
    soup = BeautifulSoup(response.text, 'html.parser')
       
    # Extract the part where to find the club ids on the website      
    all_links = soup.find_all('div', class_='large-8')
    tabelle = all_links[0].find_all('div', class_='box')
    all_zentriert = tabelle[2].find_all('td', class_='zentriert')

    # Creating an empty list
    club_ids = []

    # Collect all club ids
    for item in all_zentriert:
        item = item.find('a', class_='vereinprofil_tooltip')
        if item:
            club_ids.append(item['id'])

    # Pandas parser
    dfs = pd.read_html(response.text)
    
    # Creating an empty list
    row =[] 
    
    # Define the column titles for reindexing
    columns_titles = ['Season', 'League', '#', 'Verein', 'Verein.1',
                      'Unnamed: 3', 'G', 'U', 'V', 'Tore', '+/-', 'Pkt.']
    dfs[4]=dfs[4].reindex(columns=columns_titles)
    
    # Use known information to create a complete row
    for i in range((dfs[4].shape[0])): 
        dfs[4].iloc[i, 0] = str(season)
        dfs[4].iloc[i, 1] = league
        dfs[4].iloc[i, 3] = club_ids[i]
        CLUBLIST.append(int(club_ids[i]))
        row.append(list(dfs[4].iloc[i, :])) 
  
    # Select file
    wb = load_workbook(FILENAME)

    # Select worksheet
    ws = wb.worksheets[3]

    # Append all necessary information
    for row_data in row:
        ws.append(row_data)

    # Safe file    
    wb.save(FILENAME) 


def start_multi_scraping(leagues, seasons):
    '''
    Starts scraping with a list of leagues and a list of seasons.

            Parameters:
                    leagues (list of str): A list of club-ids
                    seasons (list of int): A list of seasons
            
            Returns:
                    Null - justs starts scraping
    '''
    
    # Use start_single_scraping with every league (in the list) ...
    # ... depending on each season (in the list)
    for league in leagues:
        for season in seasons:
            start_single_scraping(league, season)    


def label(label_type):
    '''
    Sets the name of the first line per sheet according to the label_type.

            Parameters:
                    label_type (str): A label-type

            Returns:
                    Null - stores all data in the excel file
    '''
    
    if label_type == 'table':
        # Select file
        wb = load_workbook(FILENAME)
    
        # Select worksheet
        ws = wb.worksheets[3]
    
        # Columns labeling
        column_name = ['Saison', 'Liga', 'Platz', 'Verein ID',
                       'Verein Name', 'Spieltag', 'G', 'U', 'V',
                       'Tore', 'Torverhältnis', 'Punkte'
                       ]
        ws.append(column_name)
    
        # Safe file    
        wb.save(FILENAME)


##############################################################################
##############################################################################


leagues = ['1-bundesliga', '2-bundesliga', '3-liga', 'premier-league',
           'primera-division', 'serie-a', 'ligue-1'
           ]
seasons = [2010]

label('table')
start_multi_scraping(leagues, seasons)
print(list(dict.fromkeys(CLUBLIST)))
